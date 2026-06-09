import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.load_workbook('NOMO_passion_tracker.xlsx')

# Palette
C_BLACK   = '1A1A1A'
C_CREAM   = 'FAF6EF'
C_DARK_GR = '1B4332'
C_MID_GR  = '2D6A4F'
C_WARM1   = 'F5ECD7'
C_WARM2   = 'EDE0C4'
C_ACCENT  = 'A0522D'
C_GOLD    = 'D4A017'
C_BORDER  = 'C8B89A'

thin = Side(style='thin', color=C_BORDER)
med  = Side(style='medium', color='888888')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex): return PatternFill('solid', fgColor=hex)
def font(hex, sz=10, bold=False, italic=False):
    return Font(color=hex, size=sz, bold=bold, italic=italic, name='Calibri')
def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def banner(ws, text, cols, row=1, height=34):
    ws.row_dimensions[row].height = height
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font  = font('FAF6EF', 13, bold=True)
    ws.cell(row=row, column=1).fill  = fill(C_DARK_GR)
    ws.cell(row=row, column=1).alignment = align('left', 'center')
    for ci in range(2, cols+1):
        ws.cell(row=row, column=ci).fill = fill(C_DARK_GR)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

def subtitle(ws, text, cols, row=2, height=18):
    ws.row_dimensions[row].height = height
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font  = font(C_ACCENT, 8, italic=True)
    ws.cell(row=row, column=1).fill  = fill(C_WARM1)
    ws.cell(row=row, column=1).alignment = align('center', 'center')
    for ci in range(2, cols+1):
        ws.cell(row=row, column=ci).fill = fill(C_WARM1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

def spacer(ws, row, cols, height=8):
    ws.row_dimensions[row].height = height
    for ci in range(1, cols+1):
        ws.cell(row=row, column=ci).fill = fill(C_CREAM)

def header_row(ws, headers, row, height=26):
    ws.row_dimensions[row].height = height
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci)
        cell.value = h
        cell.font  = font('FAF6EF', 9, bold=True)
        cell.fill  = fill(C_MID_GR)
        cell.alignment = align('center', 'center', wrap=True)
        cell.border = brd

# ════════════════════════════════
# 👤 MY PROFILE
# ════════════════════════════════
ws = wb['👤 My Profile']
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 4
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 28

banner(ws, 'NOMO — MULTIPASSIONATE TRACKER', 4)
subtitle(ws, 'Fill this once — your profile drives the entire tracker', 4)
spacer(ws, 3, 4)

labels = {4: 'YOUR NAME', 5: 'START DATE', 6: 'TOTAL PASSIONS'}
for r, lbl in labels.items():
    ws.row_dimensions[r].height = 26
    for ci in range(1, 5):
        ws.cell(row=r, column=ci).fill = fill(C_WARM1)
    ws.cell(row=r, column=1).value = lbl
    ws.cell(row=r, column=1).font  = font(C_DARK_GR, 10, bold=True)
    ws.cell(row=r, column=1).alignment = align('right', 'center')
    ws.cell(row=r, column=4).fill  = fill('FFFFFF')
    ws.cell(row=r, column=4).font  = font(C_BLACK, 10)
    ws.cell(row=r, column=4).alignment = align('left', 'center')
    ws.cell(row=r, column=4).border = brd

spacer(ws, 7, 4)

ws.row_dimensions[8].height = 24
ws.cell(row=8, column=1).value = 'YOUR PASSION LIST'
ws.cell(row=8, column=1).font  = font('FAF6EF', 10, bold=True)
ws.cell(row=8, column=1).fill  = fill(C_MID_GR)
ws.cell(row=8, column=1).alignment = align('left', 'center')
for ci in range(2, 5):
    ws.cell(row=8, column=ci).fill = fill(C_MID_GR)
ws.merge_cells('A8:D8')

for idx, r in enumerate(range(9, 15)):
    bg = C_WARM1 if idx % 2 == 0 else C_WARM2
    ws.row_dimensions[r].height = 24
    for ci in range(1, 5):
        ws.cell(row=r, column=ci).fill = fill(bg)
    ws.cell(row=r, column=1).value = f'Passion {idx+1}'
    ws.cell(row=r, column=1).font  = font(C_ACCENT, 9, italic=True)
    ws.cell(row=r, column=1).alignment = align('right', 'center')
    ws.cell(row=r, column=3).fill  = fill('FFFFFF')
    ws.cell(row=r, column=3).font  = font(C_DARK_GR, 10, bold=True)
    ws.cell(row=r, column=3).alignment = align('left', 'center')
    ws.cell(row=r, column=3).border = brd

# ════════════════════════════════
# 📅 DAILY LOG
# ════════════════════════════════
ws2 = wb['📅 Daily Log']
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 13
ws2.column_dimensions['B'].width = 6
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 9
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 9
ws2.column_dimensions['G'].width = 14
ws2.column_dimensions['H'].width = 9
ws2.column_dimensions['I'].width = 10
ws2.column_dimensions['J'].width = 28

banner(ws2, 'DAILY LOG — One row per day · Up to 3 passions', 10)
subtitle(ws2, 'Only 1 passion? Fill Passion 1, leave 2 & 3 blank. Score adjusts automatically.', 10)

headers2 = ['DATE','DAY','PASSION 1','YES/NO','PASSION 2','YES/NO','PASSION 3','YES/NO','ENERGY\n(1-5)',"TODAY'S WIN"]
header_row(ws2, headers2, 3)

for r in range(4, 94):
    bg = C_CREAM if r % 2 == 0 else 'FFFFFF'
    ws2.row_dimensions[r].height = 20
    for ci in range(1, 11):
        cell = ws2.cell(row=r, column=ci)
        cell.fill   = fill(bg)
        cell.font   = font(C_BLACK, 9)
        cell.border = brd
        cell.alignment = align('center', 'center')
        if ci in [4, 6, 8]:
            cell.font = font(C_DARK_GR, 9, bold=True)
        if ci == 10:
            cell.alignment = align('left', 'center')

# ════════════════════════════════
# ⭐ MY SCORE
# ════════════════════════════════
ws3 = wb['⭐ My Score']
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 10

banner(ws3, 'MY NOMO SCORE', 5)
subtitle(ws3, 'Auto-calculated from your Daily Log · Updates every time you log', 5)
spacer(ws3, 3, 5)
header_row(ws3, ['COMPONENT','WEIGHT','FORMULA','YOUR SCORE','OUT OF'], 4)

score_rows = [
    ('Streak',        '40%', 'Days with at least 1 Yes in last 15', '40'),
    ('Avg Energy',    '30%', 'Average energy rating / 5 x 30',      '30'),
    ('Wins logged',   '20%', 'Win entries in last 15 days',          '20'),
    ('Participation', '10%', 'Showed up at all (bonus)',             '10'),
]
for idx, (comp, wt, frm, out) in enumerate(score_rows):
    r = 5 + idx
    bg = C_WARM1 if idx % 2 == 0 else C_WARM2
    ws3.row_dimensions[r].height = 24
    for ci in range(1, 6):
        ws3.cell(row=r, column=ci).fill   = fill(bg)
        ws3.cell(row=r, column=ci).border = brd
        ws3.cell(row=r, column=ci).alignment = align('center', 'center')
    ws3.cell(row=r, column=1).value = comp
    ws3.cell(row=r, column=1).font  = font(C_BLACK, 9, bold=True)
    ws3.cell(row=r, column=1).alignment = align('left', 'center')
    ws3.cell(row=r, column=2).value = wt
    ws3.cell(row=r, column=2).font  = font(C_DARK_GR, 9, bold=True)
    ws3.cell(row=r, column=3).value = frm
    ws3.cell(row=r, column=3).font  = font(C_ACCENT, 8, italic=True)
    ws3.cell(row=r, column=3).alignment = align('left', 'center')
    ws3.cell(row=r, column=4).font  = font(C_DARK_GR, 11, bold=True)
    ws3.cell(row=r, column=5).value = out
    ws3.cell(row=r, column=5).font  = font(C_BLACK, 9)

# Total row
ws3.row_dimensions[9].height = 28
for ci in range(1, 6):
    ws3.cell(row=9, column=ci).fill   = fill(C_DARK_GR)
    ws3.cell(row=9, column=ci).border = brd
    ws3.cell(row=9, column=ci).alignment = align('center', 'center')
ws3.cell(row=9, column=1).value = 'NOMO SCORE'
ws3.cell(row=9, column=1).font  = font('FAF6EF', 11, bold=True)
ws3.cell(row=9, column=1).alignment = align('left', 'center')
ws3.cell(row=9, column=2).value = '100%'
ws3.cell(row=9, column=2).font  = font(C_GOLD, 11, bold=True)
ws3.cell(row=9, column=5).value = '100'
ws3.cell(row=9, column=5).font  = font(C_GOLD, 11, bold=True)

# ════════════════════════════════
# 🏆 TOP_ACHIEVERS
# ════════════════════════════════
ws4 = wb['🏆 Top_Achievers']
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 10
ws4.column_dimensions['D'].width = 13
ws4.column_dimensions['E'].width = 10
ws4.column_dimensions['F'].width = 13
ws4.column_dimensions['G'].width = 13
ws4.column_dimensions['H'].width = 10

banner(ws4, 'NOMO TOP ACHIEVERS — 15-DAY ROLLING LEADERBOARD', 8)
subtitle(ws4, 'Updated nightly · Score = Streak 40% + Energy 30% + Wins 20% + Participation 10%', 8)
spacer(ws4, 3, 8)
header_row(ws4, ['RANK','NAME','STREAK','AVG ENERGY','WINS','NOMO SCORE','PREV SCORE','CHANGE'], 4)

medal_bgs = ['FBF3D5', 'F0F0F2', 'F5EBDD']
for r in range(5, 15):
    idx = r - 5
    bg = medal_bgs[idx] if idx < 3 else (C_WARM1 if idx % 2 == 0 else 'FFFFFF')
    ws4.row_dimensions[r].height = 26
    for ci in range(1, 9):
        cell = ws4.cell(row=r, column=ci)
        cell.fill      = fill(bg)
        cell.border    = brd
        cell.alignment = align('center', 'center')
        cell.font      = font(C_BLACK, 10)
    ws4.cell(row=r, column=2).alignment = align('left', 'center')
    ws4.cell(row=r, column=2).font = font(C_BLACK, 10, bold=True)
    ws4.cell(row=r, column=6).font = font(C_DARK_GR, 11, bold=True)

wb.save('NOMO_passion_tracker.xlsx')
print('Done — NOMO_passion_tracker.xlsx styled successfully.')
